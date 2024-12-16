from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import *
import os
import pandas as pd
import jieba
from flask import Flask, request, abort

from mongodb_function import *

import json
from openpyxl import load_workbook

import time
app = Flask(__name__)
line_bot_api = LineBotApi(os.environ['CHANNEL_ACCESS_TOKEN'])
handler = WebhookHandler(os.environ['CHANNEL_SECRET'])
# 讀取 CSV 檔案
df = pd.read_csv('CSV_20241208_Base.csv')

# 載入字典
file_path = 'Real_userdict_1.txt'
jieba.load_userdict(file_path)
# 讀取 CSV 檔案
df = pd.read_csv('CSV_20241208_Base.csv')

# 載入字典
file_path2 = 'Real_userdict_2.txt'

#載入問答集
QApath = r"TestLineBot.xlsx"

# 建立一個字典來儲存詞彙和權重
dictionary_words = {}  
with open(file_path2, 'r', encoding='utf-8') as f:
    for line in f:
        word, weight = line.strip().split(' ')
        dictionary_words[word] = int(weight)  # 將權重轉換為整數

# 載入 JSON 檔案
with open('LMStudio_V1.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 提取問答配對
qa_pairs = []
for message in data['messages']:
    if len(message['versions']) > 0 and message['versions'][0]['role'] == 'user':
        question = message['versions'][0]['content'][0]['text']
    if len(message['versions']) > 1 and len(message['versions'][1]['content']) > 0:
        answer = message['versions'][1]['content'][0]['text']
    else:
    # 處理 message['versions'][1]['content'] 列表為空的情況
        answer = "找不到答案"  # 或其他預設回覆
    qa_pairs.append((question, answer))

# 建立問答知識庫
qa_dict = {question: answer for question, answer in qa_pairs}


# 建立一個字典，儲存使用者和提問次數
user_questions = {}

# 設定預設詞彙列表
special_words = { "早安", "晚安", "午安", "有問題", "嗨", "你好", "您好",  "Hello"}

# 使用者最後互動時間
user_last_interaction_time = {}

# 30 秒逾時訊息
TIMEOUT_MESSAGE = "希望這些資訊能幫助您，如有任何問題，隨時歡迎與我們聯繫。"
@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)
    write_one_data(eval(body.replace('false','False')))

    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        print("Invalid signature. Please check your channel access token/channel secret.")
        abort(400)

    return 'OK'

# 建立一個函式來處理特殊詞彙
def handle_special_words(sentence):
    for word in special_words:
        if word in sentence: # 遇到特殊詞彙就 return None
            return None  
    return 0
@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    try:
        user_id = event.source.user_id
        # 更新使用者最後互動時間
        user_last_interaction_time[user_id] = time.time()
        questionSentance = event.message.text  # 使用者啟動對話之內容
        if '@讀取' in questionSentance:
            datas = read_many_datas()
            datas_len = len(datas)
            message = TextSendMessage(text=f'資料數量，一共{datas_len}條')
            line_bot_api.reply_message(event.reply_token, message)
        elif '@查詢' in questionSentance:
            datas = col_find('events')
            message = TextSendMessage(text=str(datas))
            line_bot_api.reply_message(event.reply_token, message)
        elif '@對話紀錄' in questionSentance:
            datas = read_chat_records()
            print(type(datas))
            n = 0
            text_list = []
            for data in datas:
                if '@' in data:
                     continue
                else:
                     text_list.append(data)
                     n+=1
            data_text = '\n'.join(text_list)
            message = TextSendMessage(text=data_text[:5000])
            line_bot_api.reply_message(event.reply_token, message)
        elif '@刪除' in questionSentance:
            text = delete_all_data()
            message = TextSendMessage(text=text)
            line_bot_api.reply_message(event.reply_token, message)

    #======MongoDB操作範例======
        result = handle_special_words(questionSentance)
        if result is None:    
            noneMessage ='您好！請問您想了解哪方面的資訊呢？'
            line_bot_api.reply_message(event.reply_token,TextSendMessage(text=noneMessage))

        else:
            returnMsg = get_answer_from_excel(questionSentance,QApath)
            if returnMsg is None:
                line_bot_api.reply_message(event.reply_token,TextSendMessage(text="抱歉，我不明白您的意思"))
            else:
                
                line_bot_api.reply_message(event.reply_token,TextSendMessage(text=returnMsg))
                # 延遲幾秒後，發送後續訊息
                time.sleep(5)
                line_bot_api.push_message( event.source.user_id,TextSendMessage(text="請問您滿意這個答案嗎？ (滿意/不滿意)"))
                # 啟動計時器
                def send_timeout_message():
                    if time.time() - user_last_interaction_time.get(user_id, 0) > 30:
                        line_bot_api.push_message(user_id, TextSendMessage(text=TIMEOUT_MESSAGE))
                # 使用 Timer 延遲 30 秒執行 send_timeout_message 函式
                from threading import Timer
                timer = Timer(30, send_timeout_message)
                timer.start()
               


    except Exception as e:
        print(f"Error: {e}")
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text='發生錯誤')
        )

from openpyxl import load_workbook
def get_answer_from_excel(question, excel_file):
  """
  使用 openpyxl 讀取 Excel 檔案，確認輸入的問題是否在 A 欄的內容中，
  如果是，則回傳對應 B 欄的內容。

  Args:
    question: 輸入的問題。
    excel_file: Excel 檔案的路徑。

  Returns:
    如果問題存在於 A 欄，則返回對應 B 欄的內容，否則返回 None。
  """
  try:
    # 載入 Excel 檔案
    workbook = load_workbook(excel_file)

    # 選擇第一個工作表
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):  # 從第一行開始
        cell = row[0]
        a_value = cell.value
        if a_value == question:
            # 找到匹配的問題，返回對應 B 欄的內容
            b_value = sheet.cell(row=cell.row, column=2).value
            return b_value

    # 找不到匹配的問題
    return None

  except Exception as e:
    print(f"發生錯誤：{e}")
    return None


import os
if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
